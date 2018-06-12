from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill
from openpyxl.cell import Cell
from openpyxl.utils.cell import get_column_letter
from PIL import Image


class Converter:
    def __init__(self, cell_width=3, cell_height=10):
        self.book = Workbook()
        self._cell_width = cell_width
        self._cell_height = cell_height
        self._last_bars = 0

    def save(self, name, extension="xlsx"):
        for ws in self.book.worksheets:
            ws.sheet_view.zoomScale = 20
        self.book.save('%s.%s' % (name.split('.')[0], extension))

    def color_cell(self, x, y, rgb):
        color = '%02x%02x%02x%02x' % (255, rgb[0], rgb[1], rgb[2])
        color = color.upper()
        fill = PatternFill(start_color=color,
                           end_color=color, fill_type='solid')
        c = self.sheet.cell(row=y + 1, column=x + 1)
        c.fill = fill

    def color_current_sheet(self, image):
        self.sheet = self.book.active
        width, height = image.size
        for x in range(width):
            for y in range(height):
                rgb = image.getpixel((x, y))
                self.color_cell(x, y, rgb)
                self.print_progress(x, y, width, height)

        # resize cells to make image look better
        for x in range(1, image.size[0] + 1):
            self.sheet.column_dimensions[get_column_letter(
                x)].width = self._cell_width
        for y in range(1, image.size[1] + 1):
            self.sheet.row_dimensions[y].height = self._cell_height

    def add_image(self, image_file, compression):
        if compression <= 0:
            raise ValueError("compression should be a positive number")
        try:
            im = Image.open(image_file)
            if compression != 1:
                w = round(im.size[0] / compression)
                h = round(im.size[1] / compression)
                im = im.resize((w, h), Image.ANTIALIAS)
        except IOError:
            print("Cant load", image_file)
        else:
            self.color_current_sheet(im)

    def print_progress(self, x, y, width, height):
        percents = round(((x * height) + y) / (width * height) * 100)
        bars = percents // 10
        if bars != self._last_bars:
            self._last_bars = bars
            print("\r|" + "#" * bars + " " * (10 - bars) + "|", end='')
